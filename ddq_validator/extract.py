from __future__ import annotations

from dataclasses import dataclass
from typing import List, Optional

import openpyxl
import fitz
import re
import importlib

from .models import QuestionRow


@dataclass
class ColumnMap:
    """Column mapping for the provided DDQ Excel layout."""

    qid_col: int = 1      # A
    qtext_col: int = 2    # B
    answer_col: int = 3   # C (filled workbook)
    expected_col: int = 4 # D (reference workbook)


def norm_str(v) -> str:
    if v is None:
        return ""
    s = str(v)
    # keep newlines but trim outer whitespace
    return s.strip()


def load_questions(
    *,
    filled_path: str,
    colmap: ColumnMap | None = None,
    max_rows_per_sheet: int | None = None,
) -> List[QuestionRow]:
    """Extract questions + answers from `filled_path` only."""

    colmap = colmap or ColumnMap()
    wb_filled = openpyxl.load_workbook(filled_path)

    rows: List[QuestionRow] = []

    for sheet in wb_filled.sheetnames:
        ws_fill = wb_filled[sheet]

        max_row = ws_fill.max_row
        if max_rows_per_sheet is not None:
            max_row = min(max_row, max_rows_per_sheet)

        for r in range(1, max_row + 1):
            qid = norm_str(ws_fill.cell(r, colmap.qid_col).value)
            qtext = norm_str(ws_fill.cell(r, colmap.qtext_col).value)
            answer = norm_str(ws_fill.cell(r, colmap.answer_col).value)
            expected = norm_str(ws_fill.cell(r, colmap.expected_col).value)

            # Skip completely empty rows
            if not qtext and not qid and not answer:
                continue

            rows.append(
                QuestionRow(
                    sheet=sheet,
                    row_idx=r,
                    question_id=qid or None,
                    question_text=qtext,
                    answer_text=answer,
                    expected_text=expected,
                )
            )

    return rows


def _needs_ocr(text: str) -> bool:
    if not text:
        return True
    alnum = sum(1 for ch in text if ch.isalnum())
    return alnum < 30


def _ocr_pdf_text(path: str) -> str:
    try:
        pdf2image = importlib.import_module("pdf2image")
        pytesseract = importlib.import_module("pytesseract")
    except Exception:
        return ""

    try:
        images = pdf2image.convert_from_path(path, dpi=300)
    except Exception:
        return ""

    parts: List[str] = []
    for img in images:
        try:
            parts.append(pytesseract.image_to_string(img))
        except Exception:
            continue
    return "\n".join(parts).strip()


def _pdf_text(path: str) -> str:
    doc = fitz.open(path)
    parts: List[str] = []
    for page in doc:
        parts.append(page.get_text("text"))
    text = "\n".join(parts).strip()
    if _needs_ocr(text):
        ocr_text = _ocr_pdf_text(path)
        if ocr_text:
            return ocr_text
    return text


def _is_valid_qid(qid: str) -> bool:
    """Exclude date-like patterns (e.g. 01.10.2023, 30.09.2024) from being treated as question IDs."""
    if not qid:
        return False
    segments = qid.split(".")
    if len(segments) < 2 or len(segments) > 5:
        return False
    for seg in segments:
        if len(seg) > 2:
            return False
    return True


def _split_qid_chunks(text: str) -> List[tuple[str, str]]:
    pattern = re.compile(r"\b(\d+(?:\.\d+)+)\b")
    matches = list(pattern.finditer(text))
    chunks: List[tuple[str, str]] = []
    for i, m in enumerate(matches):
        qid = m.group(1)
        if not _is_valid_qid(qid):
            continue
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip()
        chunks.append((qid, chunk))
    return chunks


def load_questions_pdf(
    *,
    filled_path: str,
    max_rows_per_sheet: int | None = None,
) -> List[QuestionRow]:
    """Extract questions + answers from a filled PDF (best-effort, no reference needed)."""

    raw_text = _pdf_text(filled_path)
    # Keep line breaks to better separate question vs answer
    raw_text = re.sub(r"[ \t]+", " ", raw_text)

    chunks = _split_qid_chunks(raw_text)
    rows: List[QuestionRow] = []

    for idx, (qid, chunk) in enumerate(chunks, start=1):
        if max_rows_per_sheet is not None and idx > max_rows_per_sheet:
            break
        lines = [l.strip() for l in chunk.splitlines() if l.strip()]
        if not lines:
            qtext = ""
            answer = ""
        elif len(lines) == 1:
            qtext = ""
            answer = lines[0]
        else:
            qtext = lines[0]
            answer = lines[-1]

        rows.append(
            QuestionRow(
                sheet="PDF",
                row_idx=idx,
                question_id=qid,
                question_text=qtext,
                answer_text=answer,
                expected_text="",
            )
        )

    return rows


def consolidate_rows_by_qid(rows: List[QuestionRow]) -> List[QuestionRow]:
    """Group rows by (sheet, question_id), merge content. One result per question ID."""
    if not rows:
        return []

    groups: dict[tuple[str, str | None], list[QuestionRow]] = {}
    last_qid: str | None = None
    last_key: tuple[str, str | None] | None = None

    for row in rows:
        qid = (row.question_id or "").strip() or None
        if qid:
            last_qid = qid
            last_key = (row.sheet, qid)
        else:
            qid = last_qid
            last_key = (row.sheet, qid) if last_key and last_key[0] == row.sheet else last_key

        key = last_key or (row.sheet, qid)
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    consolidated: List[QuestionRow] = []
    for (sheet, qid), group in groups.items():
        if qid is None:
            continue
        qtexts = [r.question_text.strip() for r in group if r.question_text and r.question_text.strip()]
        answers = [r.answer_text.strip() for r in group if r.answer_text and r.answer_text.strip()]
        expecteds = [r.expected_text.strip() for r in group if r.expected_text and r.expected_text.strip()]
        first = group[0]
        consolidated.append(
            QuestionRow(
                sheet=sheet,
                row_idx=first.row_idx,
                question_id=qid,
                question_text=" ".join(qtexts) if qtexts else "",
                answer_text=" ".join(answers) if answers else "",
                expected_text=expecteds[0] if expecteds else "",
            )
        )
    return consolidated
